"""Write the Menu List's cost and serving weight into the ontology workbook.

``data/raw/menu_items.xlsx`` holds ``cost_per_kg`` and
``grammage_per_serving`` as XLOOKUP formulas against an external Menu List
workbook that only resolves on the author's machine. Everything the app
reads from those two columns is therefore whatever Excel last cached — and
the cache is wrong in places (a dosa at 9/kg where the list says 80/kg).

This script replaces those two columns with literal numbers taken from
``data/raw/menu_prices.xlsx``, so the dataset the app loads carries the
current figures on its own. Nothing else in the workbook is touched: only
the two columns' cells are rewritten, row by row, in place.

``src.preprocessor.price_list.apply_price_list()`` still overlays the same
list at load time, and stays the way a price refresh reaches the app — this
script is for when the values should live in the dataset too, so a
deployment without the Menu List starts from the truth rather than a stale
cache.

Usage::

    python scripts/bake_price_list.py --dry-run    # report, change nothing
    python scripts/bake_price_list.py              # write the workbook

An item the list doesn't mention keeps its cached number, written out as a
literal so no cell is left holding a formula that can't be resolved.
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

import openpyxl
import pandas as pd

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from src.preprocessor.price_list import (  # noqa: E402
    _match_key,
    load_price_list,
)

ONTOLOGY = ROOT / "data" / "raw" / "menu_items.xlsx"
PRICE_LIST = ROOT / "data" / "raw" / "menu_prices.xlsx"
COLUMNS = ("cost_per_kg", "grammage_per_serving")


def _column_indexes(sheet) -> dict:
    """1-based column index of each target column, by header name."""
    header = [cell.value for cell in sheet[1]]
    found = {}
    for name in COLUMNS:
        if name not in header:
            raise SystemExit(f"{ONTOLOGY.name}: no '{name}' column in row 1")
        found[name] = header.index(name) + 1
    found["item"] = header.index("item") + 1
    return found


def bake(dry_run: bool) -> int:
    prices = load_price_list(PRICE_LIST).set_index("key")
    # Cached values, for the rows the list has nothing to say about.
    cached = pd.read_excel(ONTOLOGY, sheet_name=0)

    book = openpyxl.load_workbook(ONTOLOGY)
    sheet = book.worksheets[0]
    at = _column_indexes(sheet)

    changed = {name: 0 for name in COLUMNS}
    unmatched = []

    for row in range(2, sheet.max_row + 1):
        item = sheet.cell(row, at["item"]).value
        if item is None or str(item).strip() == "":
            continue
        key = _match_key(item)
        listed = prices.loc[key] if key in prices.index else None
        if listed is None:
            unmatched.append(str(item))

        for name in COLUMNS:
            incoming = None if listed is None else listed[name]
            if incoming is None or pd.isna(incoming):
                # No number in the list: keep what the ontology cached, but
                # write it as a literal so the formula doesn't survive.
                incoming = cached.loc[row - 2, name]
                if pd.isna(incoming):
                    continue
            incoming = float(incoming)
            before = cached.loc[row - 2, name]
            if pd.isna(before) or round(float(before), 6) != round(incoming, 6):
                changed[name] += 1
            if not dry_run:
                sheet.cell(row, at[name]).value = incoming

    rows = sheet.max_row - 1
    print(f"{ONTOLOGY.name}: {rows} item rows")
    print(f"{PRICE_LIST.name}: {len(prices)} priced items")
    if unmatched:
        print(f"not in the price list ({len(unmatched)}), cached value kept: "
              f"{', '.join(unmatched[:10])}")
    else:
        print("every item resolved to a price-list row")
    for name in COLUMNS:
        print(f"  {name}: {changed[name]} of {rows} values change")

    if dry_run:
        print("dry run — workbook not written")
        return 0

    book.save(ONTOLOGY)
    print(f"wrote {ONTOLOGY}")
    return 0


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--dry-run", action="store_true",
                        help="report what would change without writing")
    args = parser.parse_args()
    return bake(args.dry_run)


if __name__ == "__main__":
    raise SystemExit(main())
